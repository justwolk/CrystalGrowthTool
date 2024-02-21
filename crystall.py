import json
import math
import re
import sys
import os
import pickle

from openpyxl import Workbook, load_workbook
import numpy as np
import pandas as pd
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QAction, QGraphicsScene, QGraphicsView, QMenu, \
    QMessageBox, QVBoxLayout, QToolBar, QPushButton, QLabel, QSpinBox, QGraphicsEllipseItem, QDialog, QLineEdit, \
    QDoubleSpinBox, QCheckBox, QDockWidget, QTreeWidget, QTreeWidgetItem, QWidget, QAbstractItemView, QListWidget, \
    QListWidgetItem, QShortcut
from PyQt5.QtGui import QPixmap, QPainter, QPen, QBrush, QColor, QKeySequence
from PyQt5.QtCore import Qt, QPoint, QPointF, QTimer

layers_and_points = []
fnames = []
current_layer = 0


class PointAndLayerViewer(QWidget):
    def __init__(self, layers):
        super().__init__()
        self.layers = layers
        self.initUI()

    def initUI(self):
        layout = QVBoxLayout()
        self.list_widget = QListWidget()
        self.list_widget.setSelectionMode(QAbstractItemView.SingleSelection)
        self.updateList()
        layout.addWidget(QLabel("Слои:"))
        layout.addWidget(self.list_widget)

        del_point_button = QPushButton("Удалить выбранную точку")
        del_point_button.clicked.connect(self.deletePoint)
        layout.addWidget(del_point_button)

        self.setLayout(layout)

        self.timer = QTimer(self)
        self.timer.timeout.connect(self.updateList)
        self.timer.start(1000)

    def updateList(self):
        current_row = self.list_widget.currentRow()
        self.list_widget.clear()
        for i, layer in enumerate(self.layers):
            layer_item = QListWidgetItem(f"Слой {i + 1}")
            layer_item.setFlags(Qt.ItemIsEnabled)
            self.list_widget.addItem(layer_item)
            for j, point in enumerate(layer):
                if isinstance(point, (QPointF, QPoint)):
                    self.list_widget.addItem(f"    Точка {j + 1}: ({point.x()}, {point.y()})")
        self.list_widget.setCurrentRow(current_row)

    def deletePoint(self):
        current_row = self.list_widget.currentRow()
        if current_row == -1:
            return

        layer_index, point_index = self.getLayerAndPointIndices(current_row)
        if point_index is not None:
            # Удалить точку
            del self.layers[layer_index][point_index]

            # Удалить слой если все точки удалены
            if not self.layers[layer_index]:
                del self.layers[layer_index]

            self.updateList()
        print(self.layers)


    def getLayerAndPointIndices(self, current_row):
        layer_count = len(self.layers)
        layer_index = 0
        point_index = None
        for i in range(layer_count):
            if current_row == 0:
                break
            current_row -= 1
            if current_row < len(self.layers[i]):
                point_index = current_row
                break
            current_row -= len(self.layers[i])
            layer_index += 1
        return layer_index, point_index

class SettingsDialog(QDialog):
    def __init__(self, parent=None):
        super(SettingsDialog, self).__init__(parent)
        # настройка окна настроек
        self.setWindowTitle("Настройки")

        self.layout = QVBoxLayout()

        self.browse_button = QPushButton("Выбрать файл")
        self.browse_button.clicked.connect(self.browseFile)
        self.layout.addWidget(self.browse_button)

        self.file_path = QLineEdit()
        self.layout.addWidget(self.file_path)

        self.sheet_name_label = QLabel("Название листа")
        self.layout.addWidget(self.sheet_name_label)
        self.sheet_name = QLineEdit()
        self.layout.addWidget(self.sheet_name)

        self.cell_start_label = QLabel("Записывать с ячейки №")
        self.layout.addWidget(self.cell_start_label)
        self.cell_start = QSpinBox()
        self.layout.addWidget(self.cell_start)

        self.lines_count_label = QLabel("Число линий")
        self.layout.addWidget(self.lines_count_label)
        self.lines_count = QSpinBox()
        self.layout.addWidget(self.lines_count)

        self.lines_size_label = QLabel("Толщина")
        self.layout.addWidget(self.lines_size_label)
        self.lines_size = QDoubleSpinBox()
        self.layout.addWidget(self.lines_size)
        # Inside the SettingsDialog class:
        self.draw_grid_checkbox = QCheckBox("Рисовать сетку")
        self.layout.addWidget(self.draw_grid_checkbox)

        self.save_button = QPushButton("Сохранить")
        self.save_button.clicked.connect(self.saveSettings)
        self.layout.addWidget(self.save_button)

        self.setLayout(self.layout)

        self.loadSettings()

    # методы из окна настроек

    def browseFile(self):
        fname = QFileDialog.getOpenFileName(self, 'Открыть файл', '', "Excel files (*.xlsx *.xls)")
        if fname[0]:
            self.file_path.setText(fname[0])


    #сохранение значений в json
    def saveSettings(self):
        documents_path = os.path.expanduser('~/Documents')
        settings_directory = os.path.join(documents_path, 'crystall')
        if not os.path.exists(settings_directory):
            os.makedirs(settings_directory)
        settings_path = os.path.join(settings_directory, 'settings.json')

        settings = {
            'excel_file': self.file_path.text(),
            'sheet_name': self.sheet_name.text(),
            'cell_start': self.cell_start.value(),
            'lines_count': self.lines_count.value(),
            'lines_size': self.lines_size.value(),
            'draw_grid': self.draw_grid_checkbox.isChecked(),
        }
        with open(settings_path, 'w') as f:
            json.dump(settings, f)
        self.close()

    #подгрузка значений из файла settings.json
    def loadSettings(self):
        documents_path = os.path.expanduser('~/Documents')
        settings_directory = os.path.join(documents_path, 'crystall')
        if not os.path.exists(settings_directory):
            os.makedirs(settings_directory)
        settings_path = os.path.join(settings_directory, 'settings.json')

        try:
            with open(settings_path, 'r') as f:
                settings = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            pass
            # не используется так как файл с настройками создается в главном окне
            #settings = {
                # 'excel_file': 'sample.xlsx',
                # 'sheet_name': 'Sheet1',
                # 'cell_start': 1,
                # 'lines_count': 15,
                # 'lines_size': 0.03,
            #}
        self.draw_grid_checkbox.setChecked(settings.get('draw_grid', False))
        self.file_path.setText(settings['excel_file'])
        self.sheet_name.setText(settings['sheet_name'])
        self.cell_start.setValue(settings['cell_start'])
        self.lines_count.setValue(settings['lines_count'])
        self.lines_size.setValue(settings['lines_size'])


class CustomGraphicsView(QGraphicsView):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.paint_mode = 1
        self.setRenderHint(QPainter.Antialiasing)
        self.setRenderHint(QPainter.SmoothPixmapTransform)
        self.setOptimizationFlag(QGraphicsView.DontAdjustForAntialiasing, True)
        self.setOptimizationFlag(QGraphicsView.DontSavePainterState, True)
        self.setViewportUpdateMode(QGraphicsView.FullViewportUpdate)
        self.setDragMode(QGraphicsView.ScrollHandDrag)
        self.setTransformationAnchor(QGraphicsView.AnchorUnderMouse)
        self.setResizeAnchor(QGraphicsView.AnchorUnderMouse)

        self.center_selected = False
        self.center = QPoint(0, 0)
        self.points = []
        self.layers = [[]]

    # приближение и отдаление колесиком
    def wheelEvent(self, event):
        zoom_factor = 1.15
        if event.angleDelta().y() > 0:
            self.scale(zoom_factor, zoom_factor)
        else:
            self.scale(1 / zoom_factor, 1 / zoom_factor)

    # удалить точки с картинки, и из массива
    def clear_points(self):
        for item in self.scene().items():
            if isinstance(item, QGraphicsEllipseItem):
                self.scene().removeItem(item)
        self.points = []
        self.layers[-1] = []


    #если центр или точки не выбраны, блокировка выбора точек и удаления точек
    #регулировка paintmode (рисовать точки или центр) относительно того, выбраны они или нет
    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton and self.parent().select_center_button.isChecked():
            self.paint_mode = 1
            self.center_selected = True
            self.center = self.mapToScene(event.pos())
            self.parent().count_points_button.setEnabled(True)
            self.parent().erase_points_button.setEnabled(False)
            self.update()
            self.viewport().update()

        elif event.button() == Qt.LeftButton and self.parent().count_points_button.isChecked():
            self.paint_mode = 2
            point = self.mapToScene(event.pos())
            self.points.append(point)
            self.layers[-1].append(point)
            self.parent().erase_points_button.setEnabled(True)
            print(self.layers)
            print(f"Координаты точки: ({point.x()}, {point.y()})")

            #layers_and_points.append(cleaned)
            #self.points = []
            #self.layers = [[]]
            ###layers_and_points[-1].append(QPointF(point.x(), point.y()))


            self.update()
            self.viewport().update()

        super().mousePressEvent(event)

    #пейнтэвент для рисования
    #https://zetcode.com/gui/pyqt5/painting/
    def paintEvent(self, event):
        super().paintEvent(event)

        painter = QPainter(self.viewport())
        painter.setRenderHint(QPainter.Antialiasing)
        painter.setRenderHint(QPainter.SmoothPixmapTransform)

        self.setPainterLineWidth(painter)  # Здесь устанавливаем ширину линии

        # рисуем сетку если стоит галочка
        if self.parent().settings['draw_grid']:
            self.draw_grid(painter)

        # рисуем солнышко после выбора центра
        if self.center_selected:
            self.draw_edges(painter, lines_count=self.parent().settings['lines_count'])

        # Рисуем точки если они выбраны
        self.draw_points(painter)

    def setPainterLineWidth(self, painter):
        line_width = self.parent().settings['lines_size']
        pen = painter.pen()
        pen.setWidthF(line_width)
        painter.setPen(pen)

    #рисуем круг если выбран центр
    def draw_edges(self, painter, lines_count, start_offset = 0.03):
        if self.center_selected:
            pixmap_rect = self.scene().itemsBoundingRect()
            radius = min(pixmap_rect.width(), pixmap_rect.height()) * 1

            for angle in range(0, 360, 360 // lines_count):
                start_point = QPointF(self.center.x() + start_offset * radius * math.cos(math.radians(angle)),
                                      self.center.y() + start_offset * radius * math.sin(math.radians(angle)))
                edge_point = QPointF(self.center.x() + radius * math.cos(math.radians(angle)),
                                     self.center.y() + radius * math.sin(math.radians(angle)))
                painter.drawLine(self.mapFromScene(start_point), self.mapFromScene(edge_point))

    #рисуем сетку
    def draw_grid(self, painter):
        pixmap_rect = self.scene().itemsBoundingRect()
        grid_step = 30  # одна линия через каждые 30 пикселов

        # рисуем вертикальные линии
        for x in range(int(pixmap_rect.left()), int(pixmap_rect.right()), grid_step):
            painter.drawLine(self.mapFromScene(x, pixmap_rect.top()), self.mapFromScene(x, pixmap_rect.bottom()))

        # рисуем горизонтальные линии
        for y in range(int(pixmap_rect.top()), int(pixmap_rect.bottom()), grid_step):
            painter.drawLine(self.mapFromScene(pixmap_rect.left(), y), self.mapFromScene(pixmap_rect.right(), y))

    #рисуем точки
    def draw_points(self, painter):

        painter.setPen(QPen(Qt.black, 3))
        self.setPainterLineWidth(painter)

        painter.drawEllipse(self.mapFromScene(self.center) - QPoint(3, 3), 6, 6)

        for point in self.points:
            painter.setPen(QPen(Qt.blue, 3))
            painter.drawEllipse(self.mapFromScene(point) - QPoint(3, 3), 6, 6)

    def new_layer(self):
        new_layer = []
        xx = self.layers.copy()
        # вычищаем от PyQt5.QtCore.
        cleaned = [QPointF(point.x(), point.y()) for sublist in xx for point in sublist]
        # аппендим к главному массиву
        layers_and_points.append(cleaned)
        self.points = []
        self.layers = [[]]







class ImageViewer(QMainWindow):
    def __init__(self):
        super().__init__()
        self.loadSettings()
        self.setWindowTitle('Кристаллатор')
        self.setGeometry(100, 100, 1000, 800)
        self.initUI()
        self.view.layers = layers_and_points
        self.view.layer = current_layer
        self.point_and_layer_viewer = PointAndLayerViewer(self.view.layers)
        self.point_and_layer_dock = QDockWidget("Слои и точки", self)
        self.point_and_layer_dock.setWidget(self.point_and_layer_viewer)
        self.addDockWidget(Qt.RightDockWidgetArea, self.point_and_layer_dock)

        self.shortcut = QShortcut(QKeySequence("Ctrl+O"), self)
        self.shortcut.activated.connect(self.settingsDialog)



    def uncheck_other_buttons(self, button):
        buttons = [self.select_center_button, self.count_points_button]
        for b in buttons:
            if b != button and b.isChecked():
                b.blockSignals(True)
                b.setChecked(False)
                b.blockSignals(False)

    #настраиваем главное окно
    def initUI(self):
        toolbar = QToolBar()
        self.addToolBar(toolbar)

        open_button = QPushButton("Открыть")
        open_button.clicked.connect(self.showDialog)
        open_button.setFixedSize(100, 40)
        toolbar.addWidget(open_button)

        close_button = QPushButton("Закрыть")
        close_button.clicked.connect(self.closeImage)
        close_button.setFixedSize(100, 40)
        toolbar.addWidget(close_button)

        self.select_center_button = QPushButton("Выбрать центр")
        self.select_center_button.setCheckable(True)
        self.select_center_button.setFixedSize(100, 40)
        toolbar.addWidget(self.select_center_button)

        self.count_points_button = QPushButton("Выбрать точки")
        self.count_points_button.setCheckable(True)
        self.count_points_button.setEnabled(False)
        self.count_points_button.setFixedSize(100, 40)
        toolbar.addWidget(self.count_points_button)

        #new
        self.select_center_button.toggled.connect(lambda: self.uncheck_other_buttons(self.select_center_button))
        self.count_points_button.toggled.connect(lambda: self.uncheck_other_buttons(self.count_points_button))

        self.remove_last_point_button = QPushButton("Удалить последнюю точку")
        self.remove_last_point_button.clicked.connect(self.remove_last_point)
        self.remove_last_point_button.setFixedSize(100, 40)
        toolbar.addWidget(self.remove_last_point_button)


        self.erase_points_button = QPushButton("Отчистить точки")
        self.erase_points_button.clicked.connect(self.erasePoints)
        self.erase_points_button.setFixedSize(100, 40)
        toolbar.addWidget(self.erase_points_button)


        new_layer_button = QPushButton("Новый слой")
        new_layer_button.clicked.connect(self.newLayer)
        new_layer_button.setFixedSize(100, 40)
        toolbar.addWidget(new_layer_button)

        save_button = QPushButton("Сохранить")
        save_button.setFixedSize(100, 40)
        save_button.clicked.connect(self.savePoints)
        toolbar.addWidget(save_button)

        settings_button = QPushButton("Настройки")
        settings_button.setFixedSize(100, 40)
        settings_button.clicked.connect(self.settingsDialog)
        toolbar.addWidget(settings_button)

        self.scene = QGraphicsScene()
        self.view = CustomGraphicsView(self.scene)
        self.setCentralWidget(self.view)

    def showDialog(self):
        #xfname = QFileDialog.getOpenFileName(self, 'Открыть файл', '', "Images (*.png *.xpm *.jpg *.bmp *.gif)")
        fnames, _ = QFileDialog.getOpenFileNames(self, 'Открыть файлы', '', "Images (*.png *.xpm *.jpg *.bmp *.gif)")
        print(fnames)
        print(fnames[0])

        #for fname in fnames:
        pixmap = QPixmap(fnames[0])
        self.scene.clear()
        self.view.center_selected = False
        self.view.points = []
        self.view.layers = [[]]
        self.view.center = QPoint(0, 0)
        self.view.center_selected = False
        self.count_points_button.setEnabled(False)
        self.erase_points_button.setEnabled(False)
        self.scene.addPixmap(pixmap)
        self.view.setScene(self.scene)


    def closeImage(self):
        self.scene.clear()

    def erasePoints(self):
        if self.view.layers[-1]:
            self.view.clear_points()
            self.view.update()
            self.view.viewport().update()
        else:
            QMessageBox.warning(self, "Ошибка", "В текущем слое нет точек для стирания.")

    def remove_last_point(self):
        if not self.view.points:  # пропустить если нет точек
            return

        last_point = self.view.points.pop()  # получить индекс точки

        # удалить последнюю точку из последнего слоя
        if len(self.view.layers[-1]) > 0:
            self.view.layers[-1].pop()

        # найти точку на сцене и удалить
        for item in self.view.scene().items():
            if isinstance(item, QGraphicsEllipseItem) and item.rect().center() == last_point:
                self.view.scene().removeItem(item)
                break

        self.view.update()
        self.view.viewport().update()

    def newLayer(self):
        self.view.new_layer()

    def savePoints(self):
        print(layers_and_points)
        try:
            wb = load_workbook(self.excel_file)
        except FileNotFoundError:
            wb = Workbook()

        # открываем или создаем файл
        if self.sheet_name in wb.sheetnames:
            ws = wb[self.sheet_name]
        else:
            ws = wb.create_sheet(title=self.sheet_name)

        start_row, start_col = self.cell_start, self.cell_start

        # записываем в файл
        for i, layer in enumerate(layers_and_points):
            for j, point in enumerate(layer):
                x_coord = point.x()
                y_coord = point.y()
                ws.cell(row=start_row + i * 2, column=start_col + j, value=x_coord)
                ws.cell(row=start_row + i * 2 + 1, column=start_col + j, value=y_coord)

        wb.save(self.excel_file)

    # открыть диалог настроек
    def settingsDialog(self):
        settings_dialog = SettingsDialog(self)
        settings_dialog.exec_()

    #подгрузить настройки из файла если он есть
    def loadSettings(self):
        try:
            documents_path = os.path.expanduser('~/Documents')
            settings_directory = os.path.join(documents_path, 'crystall')
            if not os.path.exists(settings_directory):
                os.makedirs(settings_directory)
            settings_path = os.path.join(settings_directory, 'settings.json')

            with open(settings_path, 'r') as f:
                settings = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            settings = {
                'excel_file': 'sample.xlsx',
                'sheet_name': 'Sheet1',
                'cell_start': 1,
                'lines_count': 15,
                'lines_size': 0.8,
                'draw_grid': False
            }
            with open(settings_path, 'w') as f:
                json.dump(settings, f)

        self.excel_file = settings['excel_file']
        self.sheet_name = settings['sheet_name']
        self.cell_start = settings['cell_start']
        self.lines_count = settings['lines_count']
        self.lines_size = settings['lines_size']
        self.settings = settings


if __name__ == '__main__':
    app = QApplication(sys.argv)
    viewer = ImageViewer()
    viewer.show()
    sys.exit(app.exec_())

