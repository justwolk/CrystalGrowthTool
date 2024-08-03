import copy
import math
import os
from PySide6.QtCore import Signal as pyqtSignal, Slot as pyqtSlot
from PySide6.QtCore import QPoint, QObject
from PIL import Image
import numpy as np
import openpyxl
from Model.crystal import CrystalImage
from PySide6 import QtCore

from Modules.find_intersection import check_intersections, generate_points
from Modules.intrpolate_line import cubic_spline_interpolation

class Controller(QObject):
    update_ui = pyqtSignal()

    def __init__(self, model):
        super().__init__()
        self._model = model
        self.mode = 0
        self.selected_grow_line_point_to_move = None
        self.grid_lines = None

    #TODO MVC pattern = set, get через controller, а не напрямую через модель

    # def set_attribute(self, name, value):
    #     if name in self._model._attributes:
    #         self._model.update_attribute(name, value)
    #         self._model.attributeChanged.emit(name, value)

    # def get_attribute(self, name):
    #     if name in self._model._attributes:
    #         return getattr(self._model, name)
    #     return None
    
    def save_project(self):
        self._model.save_to_json()
    
    @pyqtSlot(int)
    def set_mode(self, mode):
        self.mode = mode

    def clear_all_points_on_image(self):
        current_image = self._model.current_image
        crystal_object = self._model.crystal_object[current_image]
        crystal_object.clear_all_points()
        self.update_ui.emit()

    def on_reverse_ui_toggled(self):
        self._model.ui_enabled = not self._model.ui_enabled
    
    def transfer_layers_to_next_image(self):
        if not self._model.crystal_object[self._model.current_image].growth_lines:
            self._model.crystal_object[self._model.current_image].growth_lines = copy.deepcopy(self._model.crystal_object[self._model.current_image - 1].growth_lines)
            self.clear_all_points_on_image()

    @pyqtSlot(int)
    def go_to_image(self, target_image):
        total_images = len(self._model.crystal_object)
        current_image = self._model.current_image
        if 0 <= target_image <= total_images - 1:
            self._model.current_image = target_image

            if target_image - current_image == 1 and self._model.transfer_next_image == True: #если копируем
                self.transfer_layers_to_next_image()
                self.update_ui.emit()

    def go_next_image(self):
        self.go_to_image(self._model.current_image + 1)

    def go_previous_image(self):
        self.go_to_image(self._model.current_image - 1)

    def go_next_layer(self):
        self.go_to_layer(self._model.current_layer + 1)

    def go_previous_layer(self):
        self.go_to_layer(self._model.current_layer - 1)

    @pyqtSlot(int)
    def go_to_layer(self, target_layer):
        current_image = self._model.current_image
        crystal_object = self._model.crystal_object[current_image]

        if target_layer >= 0:
            if len(crystal_object.growth_lines) - 1 < target_layer:
                target_layer = len(crystal_object.growth_lines)

            self._model.update_attribute('current_layer', target_layer)
            current_layer = self._model.current_layer

            if current_layer >= len(crystal_object.growth_lines):
                crystal_object.growth_lines.append([])

    @pyqtSlot()
    def delete_layer(self): #жесткое удаление со сдвигом
        current_image = self._model.current_image
        crystal_object = self._model.crystal_object[current_image]
        target_layer = self._model.current_layer
        if target_layer >= 0 and target_layer < len(crystal_object.growth_lines):
            del crystal_object.growth_lines[target_layer]
            if self._model.current_layer > target_layer:
                self._model.update_attribute('current_layer', self._model.current_layer - 1)
            elif self._model.current_layer == target_layer:
                new_current_layer = max(len(crystal_object.growth_lines) - 1, 0)
                self._model.update_attribute('current_layer', new_current_layer)
            self.update_ui.emit()

    @pyqtSlot()
    def delete_layer_points(self): #удаление только точек
        current_image = self._model.current_image
        crystal_object = self._model.crystal_object[current_image]
        target_layer = self._model.current_layer
        if target_layer >= 0 and target_layer < len(crystal_object.growth_lines):
            crystal_object.growth_lines[target_layer] = []
        self.update_ui.emit()

    def save_to_exсel(self, pixmap):
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.create_sheet(self._model.sheet_name)
            workbook.active = sheet

            all_points = {}

            for i, crystal_obj in enumerate(self._model.crystal_object):
                for (x, y), value in crystal_obj.points.items():
                    all_points[(i, x, y)] = value

            image_count = len(self._model.crystal_object)
            all_points_array = np.empty((image_count, 100, 100), dtype=object)

            for (i, x, y), value in all_points.items():
                all_points_array[i, x, y] = value

            current_row = 0
            #TODO жесткий лимит на 199 ступеней и линий сетки
            for i in range(199):  # ступень
                if np.any(all_points_array[:, i, :]):
                    current_row += 2
                    sheet.cell(row=current_row, column=1, value=f"Ступень {i+1}")

                    for j in range(199):  # линия сетки
                        if np.any(all_points_array[:, i, j]):
                            current_row += 1
                            if self._model.save_half:
                                sheet.cell(row=current_row, column=1, value=float(j+1)/2)

                            for k in range(image_count):  # картинка
                                if all_points_array[k, i, j] is not None:
                                    x = all_points_array[k, i, j].x() / pixmap.width() * self._model.nm_value
                                    y = (pixmap.width() - all_points_array[k, i, j].y()) / pixmap.width() * self._model.nm_value

                                    sheet.cell(row=current_row, column=k * 2 + 2, value=x)
                                    sheet.cell(row=current_row, column=k * 2 + 3, value=y)

                                    sheet.cell(row=1, column=k * 2 + 3, value=str(int(k)+1))  # номер картинки

            workbook.save(self._model.excel_file)

        except Exception as e:
            print(f"Возникла ошибка: {e}")

    def reset_center_position(self, pixmap):
        self._model.center_x = int(pixmap.width()/2)
        self._model.center_y = int(pixmap.width()/2)

    @pyqtSlot(int, int)
    def apply_center_position(self, x, y):
        self._model.center_x = x
        self._model.center_y = y

    @pyqtSlot(str)
    def crop_images(self, folder_path):
        crop_box = (136, 6, 136 + 721, 6 + 721)
        new_folder_path = f"{folder_path}_cropped"
        os.makedirs(new_folder_path, exist_ok=True)

        for filename in os.listdir(folder_path):
            if filename.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.gif')):
                img_path = os.path.join(folder_path, filename)
                img = Image.open(img_path)
                cropped_img = img.crop(crop_box)
                cropped_img.save(os.path.join(new_folder_path, filename))

    @pyqtSlot(str)
    def image_folder_selected(self, folder_path):
        self._model.image_folder = folder_path
        files = os.listdir(folder_path)
        # отфильтровать и отсортировать картинки
        photos = [file for file in files if file.lower().endswith(('.jpg', '.jpeg', '.png'))]
        photos.sort()
        self._model.crystal_object = [CrystalImage(photo_path) for photo_path in photos]
        self._model.photos = photos

    @pyqtSlot()
    def generate_points_layer(self):
        growth_lines = self._model.crystal_object[self._model.current_image].growth_lines[self._model.current_layer]
        if growth_lines:
            if len(growth_lines) > 3:
                growth_lines_to_calculate = cubic_spline_interpolation(growth_lines)

                if self._model.interpolation_enabled: #TODO
                    pass 
                else:
                    pass 

                growth_lines_to_calculate = np.array([[point.x(), point.y()] for point in growth_lines_to_calculate])
                lines = self.grid_lines
                lines = np.array([[(point1.x(), point1.y()), (point2.x(), point2.y())] for point1, point2 in lines])

                #result = check_intersections(lines, growth_lines_to_calculate)
                result = generate_points(growth_lines_to_calculate, lines)

                if self._model.interpolation_enabled: #TODO
                    pass
                else:
                    print("Не работает")

                for intersection, line_idx in result:
                    intersection_point = QPoint(intersection[0], intersection[1])
                    self._model.crystal_object[self._model.current_image].set_point(self._model.current_layer, line_idx, intersection_point)
        self.update_ui.emit()

    @pyqtSlot(int)
    def generate_points_image(self):
        current_layer = self._model.current_layer
        growth_lines_count = len(self._model.crystal_object[self._model.current_image].growth_lines)
        for layer in range(growth_lines_count):
            self._model.current_layer = layer
            self.generate_points_layer()
        self.update_ui.emit()
        self._model.current_layer = current_layer

    @pyqtSlot()
    def generate_points_all(self):
        image_count = len(self._model.crystal_object)
        current_image = self._model.current_image
        for image_number in range(image_count):
            self.go_to_image(image_number)
            self.generate_points_image()
        self.update_ui.emit()
        self._model.current_image = current_image

    def process_second_image_clicked(self, event, original_size, current_size):
        click_pos = event.pos()
        #пересчитать правильную координату клика на изображение, так как размер окна под картинку != исходному размеру картинки
        newClickPos = self.recalculate_point(original_size, current_size, click_pos)
        current_image = self._model.current_image
        current_layer = self._model.current_layer
        crystal_object = self._model.crystal_object[current_image]

        if event.button() == QtCore.Qt.LeftButton:

            #удалить точку для передвижения, если выбран другой режим
            if self.selected_grow_line_point_to_move != None and self.mode != 4:
                self.selected_grow_line_point_to_move = None

            if self.mode == 1: #выбрать центр
                self.apply_center_position(newClickPos.x(), newClickPos.y())

            elif self.mode == 3: #добавить точку линии роста
                if not crystal_object.growth_lines:
                    crystal_object.growth_lines.append([newClickPos])
                else:
                    crystal_object.growth_lines[current_layer].append(newClickPos)
            
            elif self.mode == 4: #двигать точку линии роста
                    clickThreshold = 100  # пикселей
                    points = crystal_object.growth_lines[current_layer]
                    # ищем самую близкую точку для удаления
                    if self.selected_grow_line_point_to_move: #TODO
                        closestPoint = None
                        minDistance = float('inf')
                        for point in crystal_object.growth_lines[current_layer]:
                            distance = (point - self.selected_grow_line_point_to_move).manhattanLength()
                            if distance < minDistance:
                                minDistance = distance
                                closestPoint = point
                        crystal_object.growth_lines[current_layer][points.index(closestPoint)] = newClickPos
                        self.selected_grow_line_point_to_move = None
                    else:
                        self.selected_grow_line_point_to_move = newClickPos
        
            elif self.mode == 5: #удалить точку линии роста = 5
                    clickThreshold = 100  # пикселей
                    # ищем самую близкую точку для удаления
                    closestPoint = None
                    minDistance = float('inf')
                    for point in crystal_object.growth_lines[current_layer]:
                        distance = (point - newClickPos).manhattanLength()
                        if distance < minDistance:
                            minDistance = distance
                            closestPoint = point

                    if minDistance <= clickThreshold:
                        crystal_object.growth_lines[current_layer].remove(closestPoint)
    
            elif self.mode == 7: #выбрать точки роста
                    is_magnet_enabled = self._model.magnet_enabled
                    is_override_line = self._model.override_line

                    closest_line = min(self.grid_lines, key=lambda line: self.distance_to_line(newClickPos, line))
                    clicked_line_index = self.grid_lines.index(closest_line) + 1
                    print(f"Линия кликнута: {clicked_line_index}")
                    print(f"newClickPos: {newClickPos}")

                    if is_override_line == False:
                        if is_magnet_enabled == True:
                            recalculated_point = self.recalculate_point_on_line(newClickPos, clicked_line_index - 1)
                            crystal_object.set_point(current_layer, clicked_line_index - 1, recalculated_point)
                        else:
                            crystal_object.set_point(current_layer, clicked_line_index - 1, newClickPos)

                    else:
                        line_to_override = int(self.ui.textEdit_2.text()) - 1
                        if isinstance(line_to_override, int):
                            if self._model.magnet_enabled == True:
                                recalculated_point = self.recalculate_point_on_line(newClickPos, clicked_line_index - 1)
                                crystal_object.set_point(current_layer, line_to_override, recalculated_point)
                            else:
                                crystal_object.set_point(current_layer, line_to_override, newClickPos)

            elif self.mode == 10: #удалить точку
                point_list = crystal_object.get_values_by_grow_layer(current_layer)
                closest_point_index = self.find_nearest_point(newClickPos, point_list)
                closest_line = min(self.grid_lines, key=lambda line: self.distance_to_line(newClickPos, line))
                clicked_line_index = self.grid_lines.index(closest_line) 

                print(closest_point_index)
                if closest_point_index is not None:
                    crystal_object.remove_point(current_layer, clicked_line_index)            

        if event.button() == QtCore.Qt.RightButton:
            clickThreshold = 100  # пикселей
            points = crystal_object.growth_lines[current_layer]
            if self.selected_grow_line_point_to_move: #TODO
                closestPoint = None
                minDistance = float('inf')
                for point in crystal_object.growth_lines[current_layer]:
                    distance = (point - self.selected_grow_line_point_to_move).manhattanLength()
                    if distance < minDistance:
                        minDistance = distance
                        closestPoint = point
                crystal_object.growth_lines[current_layer][points.index(closestPoint)] = newClickPos
                self.selected_grow_line_point_to_move = None
            else:
                self.selected_grow_line_point_to_move = newClickPos

        self.update_ui.emit()

    def find_nearest_point(self, clicked_point, points_list):
        closest_point_index = None
        min_distance = float('inf')

        for i, point in enumerate(points_list):
            if point is not None:
                distance = abs(clicked_point.x() - point.x()) + abs(clicked_point.y() - point.y())
                if distance < min_distance:
                    min_distance = distance
                    closest_point_index = i

        return closest_point_index

    def recalculate_point(self, original_size, current_size, point):
        scaleX = original_size.width() / current_size.width()
        scaleY = original_size.height() / current_size.height()
        recalculated_point = QPoint(point.x() * scaleX, point.y() * scaleY)
        return (recalculated_point)

    def distance_to_line(self, point, line):
        start_point, end_point = line
        distance = min(math.hypot(point.x() - start_point.x(), point.y() - start_point.y()),
                       math.hypot(point.x() - end_point.x(), point.y() - end_point.y()))
        return distance

    def recalculate_point_on_line(self, click_pos, line_index):
        start_point, end_point = self.grid_lines[line_index]

        dx = end_point.x() - start_point.x()
        dy = end_point.y() - start_point.y()

        cx = click_pos.x() - start_point.x()
        cy = click_pos.y() - start_point.y()

        dot_product = cx * dx + cy * dy
        line_length_squared = dx ** 2 + dy ** 2
        t = dot_product / line_length_squared

        recalculated_point = QtCore.QPoint(start_point.x() + t * dx, start_point.y() + t * dy)

        return recalculated_point